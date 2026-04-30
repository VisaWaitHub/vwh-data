#!/usr/bin/env python3

import csv
import json
import sys
from pathlib import Path
from typing import Dict, Any

import requests
from openpyxl import load_workbook
import pdfplumber
import re


# =========================
# CONFIG
# =========================

BASE_DIR = Path(__file__).resolve().parents[1]
DOCS_DIR = BASE_DIR / "docs"
TEST_OUTPUT_PATH = DOCS_DIR / "_country_context_test.json"
LIVE_OUTPUT_PATH = DOCS_DIR / "country_context.json"

# Put your real source URLs here later.
# For the first safe manual test, we allow fallback sample mode.
RAW_ISSUANCE_XLSX_PATH = BASE_DIR / "source_data" / "raw" / "fy2024_niv_detail.xlsx"
RAW_REFUSAL_PDF_PATH = BASE_DIR / "source_data" / "raw" / "fy2025_b_refusal_rates.pdf"

ISSUANCE_SOURCE_PATH = BASE_DIR / "source_data" / "issuance_source.csv"
REFUSAL_SOURCE_PATH = BASE_DIR / "source_data" / "refusal_source.csv"

MIN_COUNTRY_COUNT = 100
REQUEST_TIMEOUT = 60

# Locked thresholds from your current system
HIGH_DEMAND_THRESHOLD = 500_000
MODERATE_DEMAND_THRESHOLD = 150_000


# =========================
# HELPERS
# =========================

def log(msg: str) -> None:
    print(msg, flush=True)


def fail(msg: str, code: int = 1) -> None:
    print(f"ERROR: {msg}", file=sys.stderr, flush=True)
    sys.exit(code)


def demand_tier(value: int) -> str:
    if value >= HIGH_DEMAND_THRESHOLD:
        return "high"
    if value >= MODERATE_DEMAND_THRESHOLD:
        return "moderate"
    return "lower"


def format_int(value: int) -> str:
    return f"{value:,}"


def format_pct(value: float) -> str:
    return f"{value:.2f}%"


def ensure_docs_dir() -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)


def safe_get_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        fail(f"Could not read existing JSON file at {path}: {e}")


# =========================
# COUNTRY MAPPING
# Expand this as needed
# =========================

COUNTRY_NAME_TO_CODE = {
    "India": "IN",
    "Brazil": "BR",
    "Nigeria": "NG",
    "Mexico": "MX",
    "China": "CN",
    "Philippines": "PH",
    "Colombia": "CO",
    "Argentina": "AR",
    "Turkey": "TR",
    "South Africa": "ZA",
    "Kenya": "KE",
    "Ghana": "GH",
    "United Kingdom": "GB",
    "France": "FR",
    "Germany": "DE",
    "South Korea": "KR",
    "United Arab Emirates": "AE",
    "Chile": "CL",
    "Peru": "PE",
    "Egypt": "EG",
    "United States": "US",
    "Canada": "CA",
    "Australia": "AU",
    "Japan": "JP",
    "Spain": "ES",
    "Italy": "IT",
    "Netherlands": "NL",
    "Sweden": "SE",
    "Switzerland": "CH",
    "Singapore": "SG",
    "Malaysia": "MY",
    "Thailand": "TH",
    "Vietnam": "VN",
    "Indonesia": "ID",
    "Pakistan": "PK",
    "Bangladesh": "BD",
    "Sri Lanka": "LK",
    "Nepal": "NP",
    "Saudi Arabia": "SA",
    "Qatar": "QA",
    "Kuwait": "KW",
    "Oman": "OM",
    "Israel": "IL",
    "Poland": "PL",
    "Czech Republic": "CZ",
    "Hungary": "HU",
    "Romania": "RO",
    "Greece": "GR",
    "Portugal": "PT",
    "Ireland": "IE",
    "Norway": "NO",
    "Denmark": "DK",
    "Finland": "FI",
    "New Zealand": "NZ",
    "Ethiopia": "ET",
    "Tanzania": "TZ",
    "Uganda": "UG",
    "Senegal": "SN",
    "Morocco": "MA",
    "Algeria": "DZ",
    "Tunisia": "TN",
    "Jordan": "JO",
    "Lebanon": "LB",
    "Iran": "IR",
    "Iraq": "IQ",
    "Bolivia": "BO",
    "Uruguay": "UY",
    "Paraguay": "PY",
    "Ecuador": "EC",
    "Venezuela": "VE",
    "Costa Rica": "CR",
    "Panama": "PA",
    "Guatemala": "GT",
    "Honduras": "HN",
    "El Salvador": "SV",
    "Nicaragua": "NI",
    "Dominican Republic": "DO",
    "Jamaica": "JM",
    "Trinidad and Tobago": "TT",
    "Bahamas": "BS",
    "Iceland": "IS",
    "Luxembourg": "LU",
    "Malta": "MT",
    "Cyprus": "CY",
    "Estonia": "EE",
    "Latvia": "LV",
    "Lithuania": "LT",
    "Slovenia": "SI",
    "Slovakia": "SK",
    "Bulgaria": "BG",
    "Croatia": "HR",
    "Serbia": "RS",
    "Ukraine": "UA",
    "Belarus": "BY",
    "Kazakhstan": "KZ",
    "Uzbekistan": "UZ",
    "Azerbaijan": "AZ",
    "Georgia": "GE",
    "Armenia": "AM",
    "Cambodia": "KH",
    "Laos": "LA",
    "Myanmar": "MM",
    "Mongolia": "MN",
    "Afghanistan": "AF",
    "Zimbabwe": "ZW",
    "Zambia": "ZM",
    "Botswana": "BW",
    "Namibia": "NA",
    "Malawi": "MW",
    "Mozambique": "MZ",
}


# =========================
# FETCH
# =========================

def fetch_text(url: str) -> str:
    log(f"Fetching text: {url}")
    r = requests.get(url, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return r.text


def fetch_bytes(url: str) -> bytes:
    log(f"Fetching binary: {url}")
    r = requests.get(url, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return r.content


# =========================
# PARSE PLACEHOLDERS
# Replace these later with real source parsers.
# For the first manual test, they can use sample data.
# =========================
def normalize_country_name(name: str) -> str:
    """
    Normalizes source names to the exact names used by your site / CSVs.
    """
    n = " ".join(str(name).strip().split())

    replacements = {
        "Great Britain and Northern Ireland": "United Kingdom",
        "United Kingdom / Great Britain and Northern Ireland": "United Kingdom",
        "Korea, South": "South Korea",
        "Republic of Korea": "South Korea",
        "UAE": "United Arab Emirates",
        "China - mainland": "China",
        "China - mainland born": "China",
        "China - Mainland born": "China",
        "China (Mainland-born)": "China",
        "China, mainland born": "China",
    }

    return replacements.get(n, n)
def get_refusal_pdf_name_variants(site_country: str) -> list[str]:
    """
    Returns possible country names as they may appear in the State Dept refusal PDF.
    """
    variants = {
        "United Kingdom": [
            "United Kingdom",
            "Great Britain and Northern Ireland",
        ],
        "South Korea": [
            "South Korea",
            "Korea, South",
            "Republic of Korea",
        ],
        "United Arab Emirates": [
            "United Arab Emirates",
            "UAE",
        ],
        "China": [
            "China",
            "China - mainland",
            "China - mainland born",
            "China, mainland born",
            "China (Mainland-born)",
        ],
    }

    return variants.get(site_country, [site_country])    
def rebuild_issuance_csv_from_excel() -> None:
    """
    Reads the official FY2024 NIV Detail Table Excel file and rebuilds:
    source_data/issuance_source.csv

    This State Department file does not have a simple "Total" column.
    Row 1 contains visa-class headers.
    Column 1 contains country / region names.
    Columns 2+ contain issuance counts by visa class.

    We calculate each country's total by summing columns 2+.
    """
    if not RAW_ISSUANCE_XLSX_PATH.exists():
        log(f"Raw issuance Excel not found at {RAW_ISSUANCE_XLSX_PATH}. Skipping rebuild.")
        return

    log(f"Reading raw issuance Excel: {RAW_ISSUANCE_XLSX_PATH}")
    wb = load_workbook(RAW_ISSUANCE_XLSX_PATH, data_only=True)
    ws = wb.active

    log(f"Workbook sheets: {wb.sheetnames}")
    log(f"Active sheet: {ws.title}")
    log(f"Rows: {ws.max_row}, Columns: {ws.max_column}")

    rows_out = []

    # Row 1 is the visa-class header row.
    # Real country rows start at row 2.
    for row_idx in range(2, ws.max_row + 1):
        country_raw = ws.cell(row=row_idx, column=1).value

        if country_raw is None:
            continue

        country_name = str(country_raw).strip()
        if not country_name:
            continue

        # Skip region/header rows like "Africa" because they have no issuance values.
        row_total = 0
        has_numeric_value = False

        for col_idx in range(2, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value

            if value is None or value == "":
                continue

            try:
                numeric_value = int(float(value))
            except Exception:
                continue

            row_total += numeric_value
            has_numeric_value = True

        if not has_numeric_value:
            continue

        normalized_country = normalize_country_name(country_name)
        code = COUNTRY_NAME_TO_CODE.get(normalized_country)

        # For now, only include countries we have mapped.
        # Later, when we load the full 111-country map, this will scale automatically.
        if not code:
            log(f"UNMAPPED COUNTRY: {normalized_country}")
            continue

        rows_out.append({
            "country_code": code,
            "country": normalized_country,
            "issuance_volume_value": row_total
        })

    if len(rows_out) < MIN_COUNTRY_COUNT:
        fail(f"Issuance Excel rebuild produced too few rows: {len(rows_out)}")

    rows_out = sorted(rows_out, key=lambda r: r["country"])

    with ISSUANCE_SOURCE_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["country_code", "country", "issuance_volume_value"]
        )
        writer.writeheader()
        writer.writerows(rows_out)

    log(f"Rebuilt issuance CSV with {len(rows_out)} rows: {ISSUANCE_SOURCE_PATH}")
def parse_issuance_data() -> Dict[str, Dict[str, Any]]:
    """
    Reads source_data/issuance_source.csv

    Expected CSV columns:
    country_code,country,issuance_volume_value
    """
    if not ISSUANCE_SOURCE_PATH.exists():
        log("issuance_source.csv not found. Using sample issuance data for safe fallback.")
        return {
            "IN": {"country": "India", "issuance_volume_value": 1200000},
            "BR": {"country": "Brazil", "issuance_volume_value": 700000},
            "NG": {"country": "Nigeria", "issuance_volume_value": 80000},
            "MX": {"country": "Mexico", "issuance_volume_value": 950000},
            "FR": {"country": "France", "issuance_volume_value": 180000},
            "DE": {"country": "Germany", "issuance_volume_value": 160000},
            "CO": {"country": "Colombia", "issuance_volume_value": 300000},
            "TR": {"country": "Turkey", "issuance_volume_value": 90000},
            "CL": {"country": "Chile", "issuance_volume_value": 60000},
            "EG": {"country": "Egypt", "issuance_volume_value": 110000},
            "GB": {"country": "United Kingdom", "issuance_volume_value": 400000},
            "AE": {"country": "United Arab Emirates", "issuance_volume_value": 170000},
            "CA": {"country": "Canada", "issuance_volume_value": 350000},
            "AU": {"country": "Australia", "issuance_volume_value": 140000},
            "JP": {"country": "Japan", "issuance_volume_value": 200000},
            "KR": {"country": "South Korea", "issuance_volume_value": 190000},
            "CN": {"country": "China", "issuance_volume_value": 500000},
            "PK": {"country": "Pakistan", "issuance_volume_value": 130000},
            "ZA": {"country": "South Africa", "issuance_volume_value": 100000},
            "AR": {"country": "Argentina", "issuance_volume_value": 125000},
        }

    import csv

    output: Dict[str, Dict[str, Any]] = {}
    with ISSUANCE_SOURCE_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = (row.get("country_code") or "").strip().upper()
            country = (row.get("country") or "").strip()
            raw_value = (row.get("issuance_volume_value") or "").strip()

            if not code or not country or not raw_value:
                continue

            output[code] = {
                "country": country,
                "issuance_volume_value": int(float(raw_value)),
            }

    return output
def rebuild_refusal_csv_from_pdf() -> None:
    """
    Reads the official FY2025 B-visa adjusted refusal rates PDF and rebuilds:
    source_data/refusal_source.csv

    Expected pattern in extracted text:
    Country name followed by a percent value.
    Example:
    India 22.04%
    """
    if not RAW_REFUSAL_PDF_PATH.exists():
        log(f"Raw refusal PDF not found at {RAW_REFUSAL_PDF_PATH}. Skipping rebuild.")
        return

    log(f"Reading raw refusal PDF: {RAW_REFUSAL_PDF_PATH}")

    text_chunks = []
    with pdfplumber.open(RAW_REFUSAL_PDF_PATH) as pdf:
        log(f"PDF pages: {len(pdf.pages)}")
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            log(f"Extracted text from refusal PDF page {page_num}: {len(text)} characters")
            text_chunks.append(text)

    full_text = "\n".join(text_chunks)

    rows_out = []

    # We use your mapped country names as the target list.
    # This keeps the first parser safe and aligned with the countries your site currently supports.
    for site_country, code in COUNTRY_NAME_TO_CODE.items():
        possible_names = get_refusal_pdf_name_variants(site_country)

        found_rate = None
        matched_name = None

        for pdf_name in possible_names:
            # Match: country name + number + %
            pattern = re.compile(
                r"(^|\n)\s*" + re.escape(pdf_name) + r"\s+([0-9]{1,2}(?:\.[0-9]+)?|100(?:\.0+)?)\s*%",
                re.IGNORECASE
            )
            match = pattern.search(full_text)

            if match:
                found_rate = float(match.group(2))
                matched_name = pdf_name
                break

        if found_rate is None:
            continue

        rows_out.append({
            "country_code": code,
            "country": site_country,
            "refusal_rate_value": found_rate
        })

        log(f"Mapped refusal rate: {site_country} ({matched_name}) = {found_rate}%")

    if len(rows_out) < MIN_COUNTRY_COUNT:
        fail(f"Refusal PDF rebuild produced too few rows: {len(rows_out)}")

    rows_out = sorted(rows_out, key=lambda r: r["country"])

    with REFUSAL_SOURCE_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["country_code", "country", "refusal_rate_value"]
        )
        writer.writeheader()
        writer.writerows(rows_out)

    log(f"Rebuilt refusal CSV with {len(rows_out)} rows: {REFUSAL_SOURCE_PATH}")
def parse_refusal_data() -> Dict[str, Dict[str, Any]]:
    """
    Reads source_data/refusal_source.csv

    Expected CSV columns:
    country_code,country,refusal_rate_value
    """
    if not REFUSAL_SOURCE_PATH.exists():
        log("refusal_source.csv not found. Using sample refusal data for safe fallback.")
        return {
            "IN": {"country": "India", "refusal_rate_value": 22.04},
            "BR": {"country": "Brazil", "refusal_rate_value": 14.87},
            "NG": {"country": "Nigeria", "refusal_rate_value": 57.00},
            "MX": {"country": "Mexico", "refusal_rate_value": 19.20},
            "FR": {"country": "France", "refusal_rate_value": 12.30},
            "DE": {"country": "Germany", "refusal_rate_value": 10.80},
            "CO": {"country": "Colombia", "refusal_rate_value": 28.40},
            "TR": {"country": "Turkey", "refusal_rate_value": 24.70},
            "CL": {"country": "Chile", "refusal_rate_value": 11.20},
            "EG": {"country": "Egypt", "refusal_rate_value": 39.60},
            "GB": {"country": "United Kingdom", "refusal_rate_value": 9.90},
            "AE": {"country": "United Arab Emirates", "refusal_rate_value": 13.40},
            "CA": {"country": "Canada", "refusal_rate_value": 8.70},
            "AU": {"country": "Australia", "refusal_rate_value": 7.80},
            "JP": {"country": "Japan", "refusal_rate_value": 6.40},
            "KR": {"country": "South Korea", "refusal_rate_value": 8.10},
            "CN": {"country": "China", "refusal_rate_value": 25.50},
            "PK": {"country": "Pakistan", "refusal_rate_value": 36.20},
            "ZA": {"country": "South Africa", "refusal_rate_value": 16.90},
            "AR": {"country": "Argentina", "refusal_rate_value": 15.10},
        }

    import csv

    output: Dict[str, Dict[str, Any]] = {}
    with REFUSAL_SOURCE_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = (row.get("country_code") or "").strip().upper()
            country = (row.get("country") or "").strip()
            raw_value = (row.get("refusal_rate_value") or "").strip()

            if not code or not country or not raw_value:
                continue

            output[code] = {
                "country": country,
                "refusal_rate_value": float(raw_value),
            }

    return output

# =========================
# MERGE + BUILD
# =========================

def build_country_context(
    issuance_map: Dict[str, Dict[str, Any]],
    refusal_map: Dict[str, Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    all_codes = sorted(set(issuance_map.keys()) | set(refusal_map.keys()))
    output: Dict[str, Dict[str, Any]] = {}

    for code in all_codes:
        issuance = issuance_map.get(code, {})
        refusal = refusal_map.get(code, {})

        country = issuance.get("country") or refusal.get("country")
        if not country:
            fail(f"Missing country name for code {code}")

        issuance_value = issuance.get("issuance_volume_value")
        refusal_value = refusal.get("refusal_rate_value")

        if issuance_value is None:
            fail(f"Missing issuance_volume_value for {code} / {country}")

        if refusal_value is None:
            fail(f"Missing refusal_rate_value for {code} / {country}")

        output[code] = {
            "country": country,
            "issuance_volume": format_int(int(issuance_value)),
            "issuance_volume_value": int(issuance_value),
            "refusal_rate": format_pct(float(refusal_value)),
            "refusal_rate_value": float(refusal_value),
            "demand_tier": demand_tier(int(issuance_value)),
        }

    return output


# =========================
# VALIDATION
# =========================

def validate_country_context(data: Dict[str, Dict[str, Any]]) -> None:
    if len(data) < MIN_COUNTRY_COUNT:
        fail(f"Country count too low: {len(data)} found, need at least {MIN_COUNTRY_COUNT}")

    required_fields = {
        "country",
        "issuance_volume",
        "issuance_volume_value",
        "refusal_rate",
        "refusal_rate_value",
        "demand_tier",
    }

    for code, item in data.items():
        missing = required_fields - set(item.keys())
        if missing:
            fail(f"{code} missing fields: {sorted(missing)}")

        if not item["country"]:
            fail(f"{code} has blank country")

        if item["issuance_volume_value"] is None:
            fail(f"{code} has null issuance_volume_value")

        if not isinstance(item["issuance_volume_value"], int):
            fail(f"{code} issuance_volume_value must be int")

        refusal = item["refusal_rate_value"]
        if refusal is None:
            fail(f"{code} has null refusal_rate_value")

        if not isinstance(refusal, (int, float)):
            fail(f"{code} refusal_rate_value must be numeric")

        if refusal < 0 or refusal > 100:
            fail(f"{code} refusal_rate_value out of range: {refusal}")

        tier = item["demand_tier"]
        if tier not in {"high", "moderate", "lower"}:
            fail(f"{code} invalid demand_tier: {tier}")

    log(f"Validation passed for {len(data)} countries.")


# =========================
# WRITE
# =========================

def write_json(path: Path, data: Dict[str, Any]) -> None:
    path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8"
    )
    log(f"Wrote {path}")


# =========================
# MAIN
# =========================

def main() -> None:
    ensure_docs_dir()

    log("Starting build_country_context.py")

    rebuild_issuance_csv_from_excel()
    rebuild_refusal_csv_from_pdf()

    issuance_map = parse_issuance_data()
    refusal_map = parse_refusal_data()

    country_context = build_country_context(issuance_map, refusal_map)
    validate_country_context(country_context)

    write_json(LIVE_OUTPUT_PATH, country_context)
    log("Live country_context.json updated successfully.")


if __name__ == "__main__":
    main()
